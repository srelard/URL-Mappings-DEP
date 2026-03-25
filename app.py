"""
URL Redirect Mapping Tool
=========================
Replaces the Excel VBA macro for mapping old henkel-adhesives.com URLs
to new next.henkel-adhesives.com URLs, with async HTTP status checks.

Run with:  streamlit run app.py
"""

import asyncio
import csv
import io
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

import aiohttp
import openpyxl
import pandas as pd
import streamlit as st

# =========================================================
# Constants
# =========================================================
CATEGORIES_CSV = Path(__file__).parent / "categories.csv"
NEW_BASE_DOMAIN = "https://next.henkel-adhesives.com"
OLD_DOMAIN_MARKER = ".com/"
HTTP_CONCURRENCY = 30
HTTP_TIMEOUT = aiohttp.ClientTimeout(total=12, connect=5)

# Country/language display names for the sitemaps we process
COUNTRY_NAMES = {
    "ae": "United Arab Emirates",
    "dz": "Algeria",
    "eg": "Egypt",
    "hn": "Honduras",
    "il": "Israel",
    "ke": "Kenya",
    "ma": "Morocco",
    "ng": "Nigeria",
    "pk": "Pakistan",
    "tn": "Tunisia",
    "tr": "Turkey",
    "tz": "Tanzania",
    "za": "South Africa",
}
LANG_NAMES = {
    "en": "English",
    "fr": "French",
    "tr": "Turkish",
    "ar": "Arabic",
}

# =========================================================
# Data Loading
# =========================================================
@st.cache_data
def load_cat_dict() -> dict[str, str]:
    """Load Product Categories CSV into {norm_key: henkel_code} dict."""
    cat_dict: dict[str, str] = {}
    with open(CATEGORIES_CSV, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            nk = row["norm_key"].strip()
            hc = row["henkel_code"].strip()
            if nk and hc and nk not in cat_dict:
                cat_dict[nk] = hc
    return cat_dict



def read_sitemap_urls(file_obj) -> list[str]:
    """Read URLs from column A of an uploaded Excel file.
    Auto-adds the homepage URL if missing."""
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    urls: list[str] = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        val = row[0]
        if val is not None:
            s = str(val).strip()
            if s.startswith("http"):
                urls.append(s)
    wb.close()

    # Auto-add homepage URL if missing
    if urls:
        country, lang = detect_country_lang(urls)
        if country and lang:
            homepage = f"https://www.henkel-adhesives.com/{country}/{lang}.html"
            if homepage not in urls:
                urls.insert(0, homepage)

    return urls


def detect_country_lang(urls: list[str]) -> tuple[str, str]:
    """Detect country/lang from a list of henkel-adhesives.com URLs."""
    for url in urls:
        p = url.lower().find(".com/")
        if p < 0:
            continue
        after = url[p + 5:]
        parts = after.split("/")
        if len(parts) >= 2:
            country = parts[0].strip().lower()
            lang = remove_html_ext(parts[1].strip()).lower()
            if country and lang and len(country) <= 3 and len(lang) <= 3:
                return country, lang
    return "", ""


# =========================================================
# String Helpers
# =========================================================
def norm_text(s: str) -> str:
    """Normalize text for category matching (mirrors VBA NormText)."""
    t = s.lower().strip()
    t = t.replace("-", "").replace(" ", "").replace("\u00a0", "")
    return t


def strip_query_hash(url: str) -> str:
    p = url.find("?")
    if p >= 0:
        url = url[:p]
    p = url.find("#")
    if p >= 0:
        url = url[:p]
    return url


def remove_trailing_slash(url: str) -> str:
    return url.rstrip("/")


def remove_html_ext(s: str) -> str:
    if s.lower().endswith(".html"):
        return s[:-5]
    return s


def contains_text(haystack: str, needle: str) -> bool:
    return needle in haystack.lower()


# =========================================================
# URL Mapping Logic
# =========================================================
def build_product_url(
    parts: list[str], base: str, fallback: str, cat_dict: dict[str, str]
) -> str:
    """Try to match URL slugs against Product Categories. Last slug first,
    then walk backwards through path segments (skip country/lang/section)."""
    last_idx = len(parts) - 1
    if last_idx < 0:
        return fallback

    # 1) Try last slug (filename without .html)
    candidate = remove_html_ext(parts[last_idx])
    norm = norm_text(candidate)
    if norm and norm in cat_dict:
        return f"{base}/products.html/producttype_{cat_dict[norm]}.html"

    # 2) Walk backwards from second-to-last to index 3
    if last_idx >= 3:
        for i in range(last_idx - 1, 2, -1):
            candidate = parts[i]
            norm = norm_text(candidate)
            if norm and norm in cat_dict:
                return f"{base}/products.html/producttype_{cat_dict[norm]}.html"

    return fallback


def build_new_url(old_url: str, cat_dict: dict[str, str]) -> str:
    """Map an old henkel-adhesives.com URL to a new next.henkel-adhesives.com URL."""
    url = strip_query_hash(old_url)
    url = remove_trailing_slash(url)

    # Extract path after .com/
    p = url.lower().find(OLD_DOMAIN_MARKER)
    if p < 0:
        return ""
    after_com = url[p + len(OLD_DOMAIN_MARKER) :]

    parts = after_com.split("/")
    if len(parts) < 2:
        return ""

    country = parts[0].strip()
    lang = remove_html_ext(parts[1].strip())  # Bug fix: strip .html from lang
    section = parts[2].strip() if len(parts) > 2 else ""

    if not country or not lang:
        return ""

    base = f"{NEW_BASE_DOMAIN}/{country}/{lang}"
    prod_generic = f"{base}/products.html/producttype_industrial-root-producttype.html"

    if not section:
        return f"{base}.html"

    if contains_text(section, "applications"):
        return f"{base}/applications.html"
    elif contains_text(section, "industries"):
        return f"{base}/industries.html"
    elif contains_text(section, "insights"):
        return f"{base}/knowledge.html"
    elif contains_text(section, "search"):
        return prod_generic
    elif contains_text(section, "services"):
        return f"{base}/support.html"
    elif contains_text(section, "spotlights"):
        return f"{base}/knowledge.html"
    elif contains_text(section, "about"):
        return f"{base}.html"
    elif contains_text(section, "product"):
        return build_product_url(parts, base, prod_generic, cat_dict)
    else:
        return f"{base}.html"


# =========================================================
# HTTP Status Labels (mirrors VBA exactly)
# =========================================================
def status_label_ow(code: int, error_msg: str = "") -> str:
    """OW check label (Module2 - follows redirects)."""
    if code <= 0:
        return f"ERROR ({error_msg})" if error_msg else "ERROR"
    if 200 <= code <= 299:
        return f"OK ({code})"
    if code in (301, 302):
        return f"Redirect OK ({code})"
    if 300 <= code <= 399:
        return f"Redirect ({code})"
    if code == 404:
        return "404 Not Found"
    if 400 <= code <= 499:
        return f"ERROR ({code})"
    if 500 <= code <= 599:
        return f"Server Error ({code})"
    return f"HTTP {code}"


def status_label_dep(code: int) -> str:
    """DEP check label (Module1 - no redirect follow)."""
    if 200 <= code <= 299:
        return "OK"
    if 300 <= code <= 399:
        return "Redirect"
    if code == 401:
        return "Unauthorized (401)"
    if code == 403:
        return "Forbidden (403)"
    if code == 404:
        return "Not Found (404)"
    if code == 408:
        return "Request Timeout (408)"
    if code == 429:
        return "Too Many Requests (429)"
    if 500 <= code <= 599:
        return f"Server Error ({code})"
    if code == -1:
        return "Request failed"
    return f"HTTP {code}"


# =========================================================
# Async HTTP Checker
# =========================================================
async def _check_one(
    session: aiohttp.ClientSession,
    url: str,
    sem: asyncio.Semaphore,
    follow_redirects: bool,
) -> tuple[str, int, str]:
    """Check a single URL. Returns (url, status_code, error_msg)."""
    async with sem:
        # Try HEAD first
        try:
            async with session.head(
                url, allow_redirects=follow_redirects, timeout=HTTP_TIMEOUT
            ) as resp:
                return (url, resp.status, "")
        except Exception:
            pass

        # Fallback to GET
        try:
            async with session.get(
                url, allow_redirects=follow_redirects, timeout=HTTP_TIMEOUT
            ) as resp:
                return (url, resp.status, "")
        except Exception as e:
            return (url, -1, str(e)[:80])


async def _check_all(
    urls: list[str],
    follow_redirects: bool,
) -> dict[str, tuple[int, str]]:
    """Check all URLs async with concurrency limit. Returns {url: (code, err)}."""
    sem = asyncio.Semaphore(HTTP_CONCURRENCY)
    connector = aiohttp.TCPConnector(limit=HTTP_CONCURRENCY, ssl=False)
    async with aiohttp.ClientSession(
        connector=connector,
        headers={"User-Agent": "Mozilla/5.0 (URL-Mapping-Tool/1.0)"},
    ) as session:
        tasks = [_check_one(session, u, sem, follow_redirects) for u in urls]
        results: dict[str, tuple[int, str]] = {}
        for coro in asyncio.as_completed(tasks):
            url, code, err = await coro
            results[url] = (code, err)
        return results


def run_checks(
    urls: list[str], follow_redirects: bool, progress_bar=None
) -> dict[str, tuple[int, str]]:
    """Sync wrapper for async HTTP checks. Processes in batches for
    thread-safe progress bar updates."""
    unique_urls = list(dict.fromkeys(urls))
    all_results: dict[str, tuple[int, str]] = {}
    batch_size = 50
    total = len(unique_urls)

    for i in range(0, total, batch_size):
        batch = unique_urls[i : i + batch_size]
        with ThreadPoolExecutor(max_workers=1) as pool:
            future = pool.submit(
                asyncio.run,
                _check_all(batch, follow_redirects),
            )
            all_results.update(future.result())

        if progress_bar:
            done = min(i + batch_size, total)
            progress_bar.progress(
                done / total,
                text=f"Checked {done}/{total} URLs...",
            )

    return all_results


# =========================================================
# Excel Output Builder
# =========================================================
def _make_excel(headers: list[str], rows: list[list]) -> bytes:
    """Create a simple single-sheet Excel from headers + rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for r, row_data in enumerate(rows, start=2):
        for col, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=col, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_mapping_excel(old_urls: list[str], new_urls: list[str]) -> bytes:
    """Build Excel with OW URL -> DEP URL mapping."""
    rows = [[old, new] for old, new in zip(old_urls, new_urls)]
    return _make_excel(["OW URL", "DEP URL"], rows)


def build_ow_errors_excel(
    old_urls: list[str], ow_statuses: dict[str, tuple[int, str]]
) -> bytes:
    """Build Excel with OW URLs that have errors (non-2xx)."""
    rows = []
    for url in old_urls:
        code, err = ow_statuses.get(url, (-1, "Not checked"))
        if not (200 <= code <= 299):
            rows.append([url, status_label_ow(code, err)])
    return _make_excel(["OW URL", "Status"], rows)


def build_dep_errors_excel(
    new_urls: list[str], dep_statuses: dict[str, tuple[int, str]]
) -> bytes:
    """Build Excel with DEP URLs that have errors (non-2xx and non-redirect)."""
    rows = []
    seen = set()
    for url in new_urls:
        if url in seen:
            continue
        seen.add(url)
        code, _err = dep_statuses.get(url, (-1, "Not checked"))
        if not (200 <= code <= 399):
            rows.append([url, status_label_dep(code)])
    return _make_excel(["DEP URL", "Status"], rows)


# =========================================================
# Streamlit UI — Single Mode
# =========================================================
def run_single_mode(cat_dict: dict[str, str]):
    """Process a single sitemap file with full workflow."""

    uploaded = st.file_uploader(
        "Drop a sitemap Excel file here",
        type=["xlsx", "xls"],
    )

    if uploaded and "old_urls" not in st.session_state:
        urls = read_sitemap_urls(uploaded)
        st.session_state["old_urls"] = urls
        country, lang = detect_country_lang(urls)
        st.session_state["country_lang"] = f"{country}-{lang}" if country else "unknown"

    if "old_urls" not in st.session_state:
        return

    old_urls = st.session_state["old_urls"]
    cl = st.session_state.get("country_lang", "unknown")

    # --- Generate (auto-run once after upload) ---
    if "new_urls" not in st.session_state:
        st.session_state["new_urls"] = [build_new_url(u, cat_dict) for u in old_urls]

    new_urls = st.session_state["new_urls"]

    # --- Compact summary ---
    st.markdown(f"**{cl.upper()}** | {len(old_urls)} URLs loaded | {len(new_urls)} mapped")

    with st.expander("Preview mapping (first 50)", expanded=False):
        st.dataframe(
            pd.DataFrame({"OW URL": old_urls[:50], "DEP URL": new_urls[:50]}),
            use_container_width=True,
        )

    # --- HTTP Checks ---
    st.divider()
    checks_done = "ow_statuses" in st.session_state and "dep_statuses" in st.session_state

    if st.button("Run HTTP Checks (OW + DEP)", type="primary", disabled=checks_done):
        bar = st.progress(0, text="Checking OW URLs...")
        st.session_state["ow_statuses"] = run_checks(
            old_urls, follow_redirects=True, progress_bar=bar
        )
        bar.progress(0.5, text="Checking DEP URLs...")
        st.session_state["dep_statuses"] = run_checks(
            new_urls, follow_redirects=False, progress_bar=bar
        )
        bar.progress(1.0, text="All checks complete!")
        for key in ("ow_errors_xlsx", "dep_errors_xlsx", "mapping_xlsx"):
            st.session_state.pop(key, None)
        st.rerun()

    if checks_done:
        ow = st.session_state["ow_statuses"]
        dep = st.session_state["dep_statuses"]

        ow_ok = sum(1 for c, _ in ow.values() if 200 <= c <= 299)
        ow_404 = sum(1 for c, _ in ow.values() if c == 404)
        dep_ok = sum(1 for c, _ in dep.values() if 200 <= c <= 299)
        dep_redir = sum(1 for c, _ in dep.values() if 300 <= c <= 399)
        dep_404 = sum(1 for c, _ in dep.values() if c == 404)

        col1, col2, col3, col4, col5 = st.columns(5)
        col1.metric("OW OK", ow_ok)
        col2.metric("OW 404", ow_404)
        col3.metric("DEP OK", dep_ok)
        col4.metric("DEP Redirect", dep_redir)
        col5.metric("DEP 404", dep_404)

        # Show OW errors inline
        ow_err_rows = []
        for url in old_urls:
            code, err = ow.get(url, (-1, ""))
            if not (200 <= code <= 299):
                ow_err_rows.append({"OW URL": url, "Status": status_label_ow(code, err)})
        if ow_err_rows:
            with st.expander(f"OW Errors ({len(ow_err_rows)})", expanded=True):
                st.dataframe(pd.DataFrame(ow_err_rows), use_container_width=True)

        # Show DEP errors inline
        dep_err_rows = []
        seen = set()
        for old, new in zip(old_urls, new_urls):
            if new in seen:
                continue
            seen.add(new)
            code, err = dep.get(new, (-1, ""))
            if not (200 <= code <= 299):
                dep_err_rows.append({
                    "OW URL": old,
                    "DEP URL": new,
                    "Status": status_label_dep(code),
                })
        if dep_err_rows:
            with st.expander(f"DEP Errors / Redirects ({len(dep_err_rows)})", expanded=True):
                st.dataframe(pd.DataFrame(dep_err_rows), use_container_width=True)

        if st.button("Re-run checks"):
            for key in ("ow_statuses", "dep_statuses", "ow_errors_xlsx", "dep_errors_xlsx", "mapping_xlsx"):
                st.session_state.pop(key, None)
            st.rerun()

    # --- Fix DEP 404 Errors ---
    if "dep_statuses" in st.session_state:
        dep = st.session_state["dep_statuses"]
        error_rows = []
        for idx, (old, new) in enumerate(zip(old_urls, new_urls)):
            code, _err = dep.get(new, (-1, ""))
            if code == 404:
                error_rows.append({
                    "index": idx, "OW URL": old,
                    "DEP URL (404)": new, "New DEP URL": new,
                })

        if error_rows:
            st.divider()
            st.subheader(f"Fix {len(error_rows)} DEP 404 Errors")
            st.caption("Edit the 'New DEP URL' column, then click Apply.")

            error_df = pd.DataFrame(error_rows)
            edited_df = st.data_editor(
                error_df[["OW URL", "DEP URL (404)", "New DEP URL"]],
                column_config={
                    "OW URL": st.column_config.TextColumn(disabled=True, width="large"),
                    "DEP URL (404)": st.column_config.TextColumn(disabled=True, width="large"),
                    "New DEP URL": st.column_config.TextColumn(width="large"),
                },
                use_container_width=True,
                num_rows="fixed",
                key="dep_404_editor",
            )

            if st.button("Apply Fixes", type="primary"):
                updated_urls = list(new_urls)
                for i, row in enumerate(error_rows):
                    new_val = edited_df.iloc[i]["New DEP URL"]
                    if new_val and new_val != row["DEP URL (404)"]:
                        updated_urls[row["index"]] = new_val
                st.session_state["new_urls"] = updated_urls
                for key in ("dep_statuses", "dep_errors_xlsx", "mapping_xlsx"):
                    st.session_state.pop(key, None)
                st.rerun()

    # --- Downloads ---
    st.divider()
    st.subheader("Downloads")

    if "mapping_xlsx" not in st.session_state:
        st.session_state["mapping_xlsx"] = build_mapping_excel(old_urls, new_urls)

    dl1, dl2, dl3 = st.columns(3)
    dl1.download_button(
        f"Mapping ({cl})",
        data=st.session_state["mapping_xlsx"],
        file_name=f"url_mapping_{cl}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

    if "ow_statuses" in st.session_state:
        if "ow_errors_xlsx" not in st.session_state:
            st.session_state["ow_errors_xlsx"] = build_ow_errors_excel(
                old_urls, st.session_state["ow_statuses"]
            )
        ow_err = sum(1 for c, _ in st.session_state["ow_statuses"].values() if not (200 <= c <= 299))
        dl2.download_button(
            f"OW Errors ({ow_err})",
            data=st.session_state["ow_errors_xlsx"],
            file_name=f"ow_errors_{cl}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        dl2.caption("Run checks first")

    if "dep_statuses" in st.session_state:
        if "dep_errors_xlsx" not in st.session_state:
            st.session_state["dep_errors_xlsx"] = build_dep_errors_excel(
                new_urls, st.session_state["dep_statuses"]
            )
        dep_err = sum(1 for c, _ in st.session_state["dep_statuses"].values() if not (200 <= c <= 399))
        dl3.download_button(
            f"DEP Errors ({dep_err})",
            data=st.session_state["dep_errors_xlsx"],
            file_name=f"dep_errors_{cl}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        dl3.caption("Run checks first")


# =========================================================
# Streamlit UI — Batch Mode
# =========================================================
def run_batch_mode(cat_dict: dict[str, str]):
    """Process multiple sitemap files at once."""

    uploaded_files = st.file_uploader(
        "Drop all sitemap Excel files here",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        key="batch_upload",
    )

    if not uploaded_files:
        return

    # --- Process all files ---
    if "batch_results" not in st.session_state:
        all_results = {}
        progress = st.progress(0, text="Processing sitemaps...")
        for i, f in enumerate(uploaded_files):
            urls = read_sitemap_urls(f)
            cl_country, cl_lang = detect_country_lang(urls)
            cl = f"{cl_country}-{cl_lang}" if cl_country else f.name
            new_urls = [build_new_url(u, cat_dict) for u in urls]
            all_results[cl] = {"old_urls": urls, "new_urls": new_urls, "filename": f.name}
            progress.progress((i + 1) / len(uploaded_files), text=f"Mapped {cl}...")
        progress.progress(1.0, text=f"All {len(uploaded_files)} sitemaps mapped!")
        st.session_state["batch_results"] = all_results
        st.rerun()

    results = st.session_state["batch_results"]

    # --- Summary table ---
    summary_rows = []
    for cl, data in results.items():
        parts = cl.split("-", 1)
        country_code = parts[0] if parts else cl
        lang_code = parts[1] if len(parts) > 1 else ""
        country_name = COUNTRY_NAMES.get(country_code, country_code.upper())
        lang_name = LANG_NAMES.get(lang_code, lang_code.upper())
        summary_rows.append({
            "Country": country_name,
            "Language": lang_name,
            "Code": cl,
            "File": data["filename"],
            "URLs": len(data["old_urls"]),
            "Mapped": len(data["new_urls"]),
            "Failed": sum(1 for u in data["new_urls"] if not u),
        })

    summary_df = pd.DataFrame(summary_rows)
    # Highlight duplicate countries (multiple languages)
    country_counts = summary_df["Country"].value_counts()
    multi_lang = country_counts[country_counts > 1].index.tolist()
    if multi_lang:
        st.info(f"Multiple languages: **{', '.join(multi_lang)}**")
    st.dataframe(summary_df, use_container_width=True)

    # --- HTTP Checks ---
    st.divider()
    batch_checked = "batch_ow_all" in st.session_state

    if st.button("Run HTTP Checks (all countries)", type="primary", disabled=batch_checked):
        all_ow_errors = []
        all_dep_errors = []
        total_countries = len(results)
        bar = st.progress(0, text="Starting checks...")

        for i, (cl, data) in enumerate(results.items()):
            bar.progress(
                i / total_countries,
                text=f"Checking {cl.upper()} ({i+1}/{total_countries})...",
            )
            ow_statuses = run_checks(data["old_urls"], follow_redirects=True)
            dep_statuses = run_checks(data["new_urls"], follow_redirects=False)

            # Collect OW errors
            for url in data["old_urls"]:
                code, err = ow_statuses.get(url, (-1, ""))
                if not (200 <= code <= 299):
                    all_ow_errors.append({
                        "Country": cl.upper(),
                        "OW URL": url,
                        "Status": status_label_ow(code, err),
                    })

            # Collect DEP errors
            seen = set()
            for old, new in zip(data["old_urls"], data["new_urls"]):
                if new in seen:
                    continue
                seen.add(new)
                code, err = dep_statuses.get(new, (-1, ""))
                if not (200 <= code <= 399):
                    all_dep_errors.append({
                        "Country": cl.upper(),
                        "OW URL": old,
                        "DEP URL": new,
                        "Status": status_label_dep(code),
                    })

            # Store per-country statuses
            results[cl]["ow_statuses"] = ow_statuses
            results[cl]["dep_statuses"] = dep_statuses

        bar.progress(1.0, text="All checks complete!")
        st.session_state["batch_results"] = results
        st.session_state["batch_ow_all"] = all_ow_errors
        st.session_state["batch_dep_all"] = all_dep_errors
        st.rerun()

    if batch_checked:
        all_ow_errors = st.session_state["batch_ow_all"]
        all_dep_errors = st.session_state["batch_dep_all"]

        col1, col2 = st.columns(2)
        col1.metric("Total OW Errors", len(all_ow_errors))
        col2.metric("Total DEP Errors", len(all_dep_errors))

        if all_ow_errors:
            with st.expander(f"All OW Errors ({len(all_ow_errors)})", expanded=True):
                st.dataframe(pd.DataFrame(all_ow_errors), use_container_width=True)

        if all_dep_errors:
            with st.expander(f"All DEP Errors ({len(all_dep_errors)})", expanded=True):
                st.dataframe(pd.DataFrame(all_dep_errors), use_container_width=True)

        if st.button("Re-run all checks"):
            for key in ("batch_ow_all", "batch_dep_all", "batch_ow_xlsx", "batch_dep_xlsx"):
                st.session_state.pop(key, None)
            # Clear per-country statuses
            for data in st.session_state["batch_results"].values():
                data.pop("ow_statuses", None)
                data.pop("dep_statuses", None)
            st.rerun()

    # --- Downloads ---
    st.divider()
    st.subheader("Downloads")

    # Per-country mapping downloads
    st.caption("Mappings per country:")
    dl_cols = st.columns(min(len(results), 4))
    for i, (cl, data) in enumerate(results.items()):
        col = dl_cols[i % len(dl_cols)]
        xlsx = build_mapping_excel(data["old_urls"], data["new_urls"])
        col.download_button(
            f"Mapping ({cl})",
            data=xlsx,
            file_name=f"url_mapping_{cl}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"dl_map_{cl}",
        )

    # Combined error downloads
    if batch_checked:
        st.caption("Combined error reports:")
        dl_err1, dl_err2 = st.columns(2)

        if all_ow_errors:
            ow_rows = [[r["Country"], r["OW URL"], r["Status"]] for r in all_ow_errors]
            ow_xlsx = _make_excel(["Country", "OW URL", "Status"], ow_rows)
            dl_err1.download_button(
                f"All OW Errors ({len(all_ow_errors)})",
                data=ow_xlsx,
                file_name="ow_errors_all_countries.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            dl_err1.caption("No OW errors")

        if all_dep_errors:
            dep_rows = [[r["Country"], r["OW URL"], r["DEP URL"], r["Status"]] for r in all_dep_errors]
            dep_xlsx = _make_excel(["Country", "OW URL", "DEP URL", "Status"], dep_rows)
            dl_err2.download_button(
                f"All DEP Errors ({len(all_dep_errors)})",
                data=dep_xlsx,
                file_name="dep_errors_all_countries.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            dl_err2.caption("No DEP errors")


# =========================================================
# Main
# =========================================================
def main():
    st.set_page_config(page_title="URL Redirect Mapping Tool", layout="wide")
    st.title("URL Redirect Mapping Tool")
    st.caption("Henkel Adhesives Technologies — OneWeb to DEP URL Migration")

    st.markdown("""
**Purpose:** Automatically maps old OneWeb (OW) URLs to new Digital Experience Platform (DEP) URLs
based on URL section rules and product category matching.

**How to use:**
1. Upload a sitemap export (`.xlsx`, column A = URLs) — the homepage URL is added automatically if missing
2. URL mappings are generated instantly after upload
3. Click **Run HTTP Checks** to verify both OW and DEP URLs
4. Fix any DEP 404 errors directly in the table below
5. Download the final mapping and error reports
""")

    st.divider()

    # Load categories
    try:
        cat_dict = load_cat_dict()
    except FileNotFoundError:
        st.error(f"Categories file not found: {CATEGORIES_CSV}")
        st.stop()

    # Sidebar
    st.sidebar.header("Info")
    st.sidebar.metric("Product Categories", len(cat_dict))
    if st.sidebar.button("Reset"):
        st.session_state.clear()
        st.rerun()

    # Mode toggle
    mode = st.toggle("Batch Mode (multiple files)", value=False)

    if mode:
        run_batch_mode(cat_dict)
    else:
        run_single_mode(cat_dict)


if __name__ == "__main__":
    main()
